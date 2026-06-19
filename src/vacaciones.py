import openpyxl
from datetime import datetime
from pathlib import Path

# ========== CONSTANTES ==========
ARCHIVO_EMPLEADOS = "data/empleados.xlsx"
ARCHIVO_SOLICITUDES = "data/solicitudes.xlsx"
FORMATO_FECHA = "%d/%m/%Y"
PASS_ADMIN = "admin123"

CAMPOS_EMPLEADOS = ["DNI", "Nombre", "Dias_Disponibles"]
CAMPOS_SOLICITUDES = ["ID", "DNI", "Nombre", "Fecha_Inicio", "Fecha_Fin", "Dias_Solicitados", "Estado"]

# ========== FUNCIONES DE ARCHIVOS ==========

def crear_archivo_si_no_existe(ruta, encabezados):
    if not Path(ruta).exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(encabezados)
        wb.save(ruta)
        print(f"Archivo '{ruta}' creado con encabezados.")

def cargar_empleados():
    crear_archivo_si_no_existe(ARCHIVO_EMPLEADOS, CAMPOS_EMPLEADOS)
    wb = openpyxl.load_workbook(ARCHIVO_EMPLEADOS)
    ws = wb.active
    empleados = {}
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[0] and fila[1] is not None:
            empleados[str(fila[0]).strip()] = {
                "nombre": str(fila[1]).strip(),
                "dias": int(fila[2]) if fila[2] else 0
            }
    return empleados

def guardar_empleados(empleados):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(CAMPOS_EMPLEADOS)
    for dni, datos in empleados.items():
        ws.append([dni, datos["nombre"], datos["dias"]])
    wb.save(ARCHIVO_EMPLEADOS)

def obtener_solicitudes(filtro_dni=None):
    crear_archivo_si_no_existe(ARCHIVO_SOLICITUDES, CAMPOS_SOLICITUDES)
    wb = openpyxl.load_workbook(ARCHIVO_SOLICITUDES)
    ws = wb.active
    solicitudes = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[0] is not None:
            solicitud = {
                "id": int(fila[0]),
                "dni": str(fila[1]).strip(),
                "nombre": str(fila[2]).strip(),
                "fecha_ini": fila[3],
                "fecha_fin": fila[4],
                "dias": int(fila[5]) if fila[5] else 0,
                "estado": str(fila[6]).strip() if fila[6] else "PENDIENTE"
            }
            if filtro_dni is None or solicitud["dni"] == filtro_dni:
                solicitudes.append(solicitud)
    return solicitudes

def registrar_solicitud(dni, nombre, fecha_ini, fecha_fin, dias, estado):
    crear_archivo_si_no_existe(ARCHIVO_SOLICITUDES, CAMPOS_SOLICITUDES)
    wb = openpyxl.load_workbook(ARCHIVO_SOLICITUDES)
    ws = wb.active
    ids = [int(fila[0]) for fila in ws.iter_rows(min_row=2, values_only=True) if fila[0] is not None]
    nuevo_id = max(ids) + 1 if ids else 1
    ws.append([
        nuevo_id,
        dni,
        nombre,
        fecha_ini.strftime(FORMATO_FECHA),
        fecha_fin.strftime(FORMATO_FECHA),
        dias,
        estado
    ])
    wb.save(ARCHIVO_SOLICITUDES)
    return nuevo_id

def actualizar_estado_solicitud(id_solicitud, nuevo_estado):
    wb = openpyxl.load_workbook(ARCHIVO_SOLICITUDES)
    ws = wb.active
    for fila in ws.iter_rows(min_row=2):
        if int(fila[0].value) == id_solicitud:
            fila[6].value = nuevo_estado
            wb.save(ARCHIVO_SOLICITUDES)
            return True
    return False

def obtener_solicitud_por_id(id_solicitud):
    solicitudes = obtener_solicitudes()
    for s in solicitudes:
        if s["id"] == id_solicitud:
            return s
    return None

# ========== VALIDACIONES ==========

def validar_dni(dni: str) -> bool:
    return dni.isdigit() and len(dni) == 8

def parse_fecha(texto: str):
    try:
        return datetime.strptime(texto, FORMATO_FECHA)
    except ValueError:
        return None

def calcular_dias(fecha_ini, fecha_fin) -> int:
    return (fecha_fin - fecha_ini).days

def saldo_suficiente(dias_disponibles, dias_solicitados) -> bool:
    return dias_disponibles >= dias_solicitados

# ========== INTERACCION ==========

def pedir_dni():
    while True:
        dni = input("Ingrese su DNI (8 digitos): ").strip()
        if validar_dni(dni):
            return dni
        print("DNI invalido. Debe tener 8 digitos numericos.")

def pedir_fecha(mensaje):
    while True:
        texto = input(f"{mensaje} (dd/mm/aaaa): ").strip()
        fecha = parse_fecha(texto)
        if fecha:
            return fecha
        print("Formato incorrecto. Use dd/mm/aaaa")

def mostrar_solicitudes(lista):
    if not lista:
        print("No hay solicitudes para mostrar.")
        return
    print("\nSOLICITUDES")
    print("-" * 100)
    print(f"{'ID':<5} {'DNI':<12} {'Nombre':<20} {'Inicio':<12} {'Fin':<12} {'Dias':<6} {'Estado':<12}")
    print("-" * 100)
    for s in lista:
        print(f"{s['id']:<5} {s['dni']:<12} {s['nombre']:<20} {s['fecha_ini']:<12} {s['fecha_fin']:<12} {s['dias']:<6} {s['estado']:<12}")
    print("-" * 100)

# ========== MENU EMPLEADO ==========

def menu_empleado(dni, empleados):
    empleado = empleados[dni]
    nombre = empleado["nombre"]
    saldo = empleado["dias"]
    
    while True:
        print(f"\nHola {nombre} (Saldo: {saldo} dias)")
        print("1. Solicitar vacaciones")
        print("2. Ver mis solicitudes")
        print("3. Volver al menu principal")
        opcion = input("Seleccione: ").strip()
        
        if opcion == "1":
            print("\nNUEVA SOLICITUD")
            fecha_ini = pedir_fecha("Fecha de inicio")
            fecha_fin = pedir_fecha("Fecha de fin")
            if fecha_fin < fecha_ini:
                print("La fecha fin no puede ser anterior a la fecha inicio.")
                continue
            dias_solicitados = calcular_dias(fecha_ini, fecha_fin)
            if dias_solicitados <= 0:
                print("Debe solicitar al menos 1 dia.")
                continue
            
            if not saldo_suficiente(saldo, dias_solicitados):
                print(f"Saldo insuficiente (tiene {saldo}, solicita {dias_solicitados}).")
                print("La solicitud se registrara como RECHAZADA automaticamente.")
                estado = "RECHAZADA"
            else:
                estado = "PENDIENTE"
                print("Saldo suficiente. Solicitud registrada como PENDIENTE de aprobacion.")
            
            nuevo_id = registrar_solicitud(dni, nombre, fecha_ini, fecha_fin, dias_solicitados, estado)
            print(f"Solicitud registrada con ID {nuevo_id} - Estado: {estado}")
            
        elif opcion == "2":
            mis_solicitudes = obtener_solicitudes(filtro_dni=dni)
            mostrar_solicitudes(mis_solicitudes)
            
        elif opcion == "3":
            break
        else:
            print("Opcion invalida.")

# ========== MENU ADMINISTRADOR ==========

def menu_admin():
    while True:
        print("\nMENU ADMINISTRADOR")
        print("1. Ver solicitudes pendientes (aprobar/rechazar)")
        print("2. Ver todas las solicitudes")
        print("3. Volver al menu principal")
        opcion = input("Seleccione: ").strip()
        
        if opcion == "1":
            pendientes = [s for s in obtener_solicitudes() if s["estado"] == "PENDIENTE"]
            if not pendientes:
                print("No hay solicitudes pendientes.")
                continue
            mostrar_solicitudes(pendientes)
            
            try:
                id_seleccionado = int(input("\nIngrese el ID de la solicitud a gestionar (0 para salir): "))
                if id_seleccionado == 0:
                    continue
            except ValueError:
                print("ID invalido.")
                continue
            
            solicitud = obtener_solicitud_por_id(id_seleccionado)
            if not solicitud or solicitud["estado"] != "PENDIENTE":
                print("Solicitud no encontrada o no esta pendiente.")
                continue
            
            print(f"\nSolicitud #{solicitud['id']} - {solicitud['nombre']} (DNI: {solicitud['dni']})")
            print(f"Fechas: {solicitud['fecha_ini']} -> {solicitud['fecha_fin']} (Dias: {solicitud['dias']})")
            accion = input("Aprobar (a) o Rechazar (r)? ").strip().lower()
            if accion not in ["a", "r"]:
                print("Opcion invalida.")
                continue
            
            if accion == "a":
                empleados = cargar_empleados()
                dni = solicitud["dni"]
                if dni not in empleados:
                    print("Empleado no encontrado en la base de datos.")
                    continue
                if empleados[dni]["dias"] < solicitud["dias"]:
                    print("Error: el empleado ya no tiene suficiente saldo (posible inconsistencia).")
                    continue
                empleados[dni]["dias"] -= solicitud["dias"]
                guardar_empleados(empleados)
                actualizar_estado_solicitud(id_seleccionado, "APROBADA")
                print(f"Solicitud APROBADA. Se descontaron {solicitud['dias']} dias. Nuevo saldo: {empleados[dni]['dias']}")
            else:
                actualizar_estado_solicitud(id_seleccionado, "RECHAZADA")
                print("Solicitud RECHAZADA.")
                
        elif opcion == "2":
            todas = obtener_solicitudes()
            mostrar_solicitudes(todas)
            
        elif opcion == "3":
            break
        else:
            print("Opcion invalida.")

# ========== PROGRAMA PRINCIPAL ==========

def main():
    while True:
        print("\n" + "="*50)
        print("SISTEMA DE VACACIONES")
        print("="*50)
        print("1. Empleado (ingresar DNI)")
        print("2. Administrador")
        print("3. Salir")
        print("="*50)
        opcion = input("Seleccione: ").strip()
        
        if opcion == "1":
            empleados = cargar_empleados()
            if not empleados:
                print("No hay empleados cargados.")
                continue
            dni = pedir_dni()
            if dni not in empleados:
                print("DNI no registrado.")
                continue
            menu_empleado(dni, empleados)
            
        elif opcion == "2":
            clave = input("Ingrese contrasena de administrador: ").strip()
            if clave != PASS_ADMIN:
                print("Contrasena incorrecta.")
                continue
            menu_admin()
            
        elif opcion == "3":
            print("Hasta luego.")
            break
        else:
            print("Opcion invalida.")

if __name__ == "__main__":
    main()